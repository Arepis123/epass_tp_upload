import os
import re
import sys
import time
import threading
import ctypes
import configparser
import subprocess
import ftplib
import requests
import pymysql
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
from datetime import datetime

# Fix blurry/low-res UI on Windows high-DPI displays
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# Set AppUserModelID so Windows taskbar shows favicon.ico instead of Python's icon
try:
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("clab.epass.tp.upload")
except Exception:
    pass

# Determine base directory (works both as .py and as a PyInstaller .exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def resource_path(filename):
    """Return correct path for bundled resources in a PyInstaller EXE."""
    base = getattr(sys, '_MEIPASS', BASE_DIR)
    return os.path.join(base, filename)


REPORT_DIR      = os.path.join(BASE_DIR, "Report")
LOG_DIR         = os.path.join(BASE_DIR, "Log")
ARCHIVE_DIR     = os.path.join(BASE_DIR, "Archive")
CONFIG_PATH     = os.path.join(BASE_DIR, "config.ini")
LOG_FILE_PATH   = os.path.join(LOG_DIR, "upload.log")

INVALID_CHARS = re.compile(r'[\\/:*?"<>|]')


def sanitize(name: str) -> str:
    return INVALID_CHARS.sub("", name).strip()


# ─── Logging ──────────────────────────────────────────────────────────────────

def _file_log(msg):
    """Append a timestamped message to upload.log."""
    os.makedirs(LOG_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")


# ─── Config ───────────────────────────────────────────────────────────────────

_CONFIG_TEMPLATE = """\
[LoginDB]
host     = 127.0.0.1
port     = 3306
database =
username = root
password =

[DB]
host     = 127.0.0.1
port     = 3306
database = workers
username = root
password =

[FTP]
host       = ftp.example.com
port       = 21
username   = ftp_user
password   = ftp_password
remote_dir = /attach

[API]
base_url = https://example.com/eclabapi

[Schedule]
epass_folder =
gc_folder    =
run_time     = 08:00

[Email]
smtp_host    = smtp.gmail.com
smtp_port    = 587
username     =
password     =
sender_name  = ePass TP Upload
sender_email =
use_tls      = true
"""


def _ensure_config():
    if not os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "w") as f:
            f.write(_CONFIG_TEMPLATE)


def _read_config():
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(
            f"config.ini not found at:\n{CONFIG_PATH}\n\nPlease fill in your credentials."
        )
    cfg = configparser.ConfigParser()
    cfg.read(CONFIG_PATH, encoding="utf-8")
    return cfg


def load_login_db_config():
    cfg = _read_config()
    sec = cfg["LoginDB"]
    return {
        "host":     sec.get("host",     "127.0.0.1").strip(),
        "port":     int(sec.get("port", "3306").strip()),
        "database": sec.get("database", "").strip(),
        "username": sec.get("username", "root").strip(),
        "password": sec.get("password", "").strip(),
    }


def load_db_config():
    cfg = _read_config()
    sec = cfg["DB"]
    return {
        "host":     sec.get("host", "127.0.0.1").strip(),
        "port":     int(sec.get("port", "3310").strip()),
        "database": sec.get("database", "workers").strip(),
        "username": sec.get("username", "root").strip(),
        "password": sec.get("password", "").strip(),
    }


def load_ftp_config():
    cfg = _read_config()
    sec = cfg["FTP"]
    return {
        "host":       sec["host"].strip(),
        "port":       int(sec.get("port", "21").strip()),
        "username":   sec["username"].strip(),
        "password":   sec["password"].strip(),
        "remote_dir": sec["remote_dir"].strip(),
    }


def load_api_config():
    cfg = _read_config()
    base = cfg["API"]["base_url"].strip().rstrip("/")
    return {
        "epass_endpoint": f"{base}/api_epass_upd.php",
        "gc_endpoint":    f"{base}/api_tp_upd.php",
        "list_endpoint":  f"{base}/api_epass.php",
    }


def load_schedule_config():
    cfg = _read_config()
    sec = cfg["Schedule"] if "Schedule" in cfg else {}
    return {
        "epass_folder": sec.get("epass_folder", "").strip(),
        "gc_folder":    sec.get("gc_folder",    "").strip(),
        "run_time":     sec.get("run_time",     "08:00").strip(),
    }


def load_email_config():
    cfg = _read_config()
    sec = cfg["Email"] if "Email" in cfg else {}
    return {
        "smtp_host":    sec.get("smtp_host",    "smtp.gmail.com").strip(),
        "smtp_port":    int(sec.get("smtp_port", "587").strip()),
        "username":     sec.get("username",     "").strip(),
        "password":     sec.get("password",     "").strip(),
        "sender_name":  sec.get("sender_name",  "ePass TP Upload").strip(),
        "sender_email": sec.get("sender_email", "").strip(),
        "use_tls":      sec.get("use_tls",      "true").strip().lower() == "true",
    }


def save_schedule_config(epass_folder, gc_folder, run_time):
    cfg = _read_config()
    if "Schedule" not in cfg:
        cfg["Schedule"] = {}
    cfg["Schedule"]["epass_folder"] = epass_folder
    cfg["Schedule"]["gc_folder"]    = gc_folder
    cfg["Schedule"]["run_time"]     = run_time
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        cfg.write(f)


# ─── PDF Extraction ───────────────────────────────────────────────────────────

def extract_epass_info(text: str):
    """
    Extract (passport_no, name, nationality, vp_no, epass_expiry) from ePass PDF text.

    Two PDF layouts exist:
      Format A — value on same line as label:
        Passport No : C9517423
        Nationality :              (empty — value comes after MALE/FEMALE)

      Format B — label line is empty, values appear after gender anchor:
        Passport No :              (empty)
        Nationality :              (empty)
        ...
        MALE
        E5399922                   <- passport
        INDONESIA                  <- nationality
    """
    passport_no  = None
    name         = None
    nationality  = None
    vp_no        = None
    epass_expiry = None

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    # Name — try label line first (Format A)
    m = re.search(r"Name[ \t]*:[ \t]*(.+)", text)
    if m:
        name = m.group(1).strip()

    # Try to get passport from the label line using horizontal whitespace only
    # ([ \t]* instead of \s* so we never cross a newline)
    m = re.search(r"Passport[ \t]*No[ \t]*:[ \t]*(\S+)", text)
    passport_from_label = m.group(1).strip() if m else None

    # MALE/FEMALE anchor — supplies passport (Format B) and nationality (both formats)
    for i, line in enumerate(lines):
        if line in ("MALE", "FEMALE"):
            remaining = [l for l in lines[i + 1:] if "<" not in l]
            if passport_from_label:
                # Format A: passport already found; first remaining is nationality
                nationality = remaining[0] if remaining else None
            else:
                # Format B: passport and nationality both come from here
                if len(remaining) >= 2:
                    passport_no = remaining[0]
                    nationality = remaining[1]
                elif len(remaining) == 1:
                    passport_no = remaining[0]

            # Format B name fallback: walk backwards from the gender line,
            # skip label lines (contain ':') and lines that start with a digit
            if not name:
                for prev in reversed(lines[:i]):
                    if prev and ":" not in prev and not re.match(r'^\d', prev):
                        name = prev
                        break
            break

    if passport_from_label:
        passport_no = passport_from_label

    m = re.search(r"VP No\s*:\s*(\S+)", text)
    if m:
        vp_no = m.group(1).strip()

    m = re.search(r"until\s+(\d{1,2}\s+\w+\s+\d{4})", text)
    if m:
        epass_expiry = m.group(1).strip()

    return passport_no, name, nationality, vp_no, epass_expiry


def extract_gc_info(text: str):
    """
    Extract (passport_no, name, nationality, gc_expiry) from GC (Green Card) PDF text.

    Labels are Malay:
        No. K.P./ No. Pasport/ No. Dokumen E5399922
        Nama Personel NARDI
        Warganegara INDONESIA
        Tarikh Tamat Pendaftaran Personel 12/03/2027
    """
    passport_no = None
    name        = None
    nationality = None
    gc_expiry   = None

    m = re.search(r"No\.\s*K\.P\..*?No\.\s*Dokumen\s+(\S+)", text)
    if m:
        passport_no = m.group(1).strip()

    m = re.search(r"Nama Personel\s+(.+)", text)
    if m:
        name = m.group(1).strip()

    m = re.search(r"Warganegara\s+(.+)", text)
    if m:
        nationality = m.group(1).strip()

    m = re.search(r"Tarikh Tamat Pendaftaran Personel\s+(\S+)", text)
    if m:
        gc_expiry = m.group(1).strip()

    return passport_no, name, nationality, gc_expiry


# ─── Database ─────────────────────────────────────────────────────────────────

def db_lookup_contractor(passport_no: str, db_cfg: dict):
    """
    Query workers table for wkr_currentemp (contractor ID) by passport number.
    Returns contractor_id string or None if not found.
    """
    conn = pymysql.connect(
        host=db_cfg["host"],
        port=db_cfg["port"],
        user=db_cfg["username"],
        password=db_cfg["password"],
        database=db_cfg["database"],
        connect_timeout=10,
        charset="utf8mb4",
    )
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT wkr_currentemp FROM workers WHERE wkr_passno = %s LIMIT 1",
                (passport_no,)
            )
            row = cur.fetchone()
            return row[0] if row else None
    finally:
        conn.close()


# ─── FTP ──────────────────────────────────────────────────────────────────────

def _ftp_makedirs(ftp, remote_dir, log_fn=None):
    """Recursively create remote directories, logging each step."""
    parts = remote_dir.strip("/").split("/")
    path = ""
    for part in parts:
        path = f"{path}/{part}"
        try:
            ftp.mkd(path)
            if log_fn:
                log_fn(f"  [FTP] Created dir: {path}\n")
        except ftplib.error_perm as e:
            code = str(e)[:3]
            if code == "550":
                # 550 = already exists, safe to continue
                if log_fn:
                    log_fn(f"  [FTP] Dir exists:  {path}\n")
            else:
                if log_fn:
                    log_fn(f"  [FTP] mkdir error ({code}): {path} — {e}\n")
                raise


def ftp_upload_file(ftp, local_path, remote_path, log_fn):
    """Upload a single file to FTP with verbose logging."""
    size = os.path.getsize(local_path)
    log_fn(f"  [FTP] Uploading ({size:,} bytes): {os.path.basename(local_path)}\n")
    log_fn(f"  [FTP] Remote path: {remote_path}\n")
    with open(local_path, "rb") as f:
        ftp.storbinary(f"STOR {remote_path}", f)
    log_fn(f"  [FTP] Upload OK → {remote_path}\n")
    return remote_path


# ─── API ──────────────────────────────────────────────────────────────────────

def fetch_listing(list_endpoint, log_fn):
    """
    GET api_epass.php and return {passport_no: {needs_epass, needs_gc, clab_id}} dict.
    Fields ePASS / Green_Card being null means that document is still missing.
    Returns None if the request fails (caller falls back to DB lookup).
    """
    try:
        resp = requests.get(list_endpoint, timeout=30)
        data = resp.json()
        if isinstance(data, dict):
            data = [data]
        listing = {}
        for item in data:
            pp = (item.get("Passport_No") or "").strip()
            if not pp:
                continue
            listing[pp] = {
                "needs_epass": item.get("ePASS")       is None,
                "needs_gc":    item.get("Green_Card")   is None,
                "clab_id":     (item.get("CLAB_ID")     or "").strip(),
                "worker_name": (item.get("Worker_Name") or "").strip(),
            }
        log_fn(f"Listing API: {len(listing)} worker(s) with pending documents.\n\n")
        return listing
    except Exception as e:
        log_fn(f"  WARNING: Listing API unavailable ({e}) — falling back to DB lookup.\n\n")
        return None


def notify_api(url, params, log_fn):
    """GET notification to a single endpoint. Returns (success_bool, message, parsed_json_or_None)."""
    resp = requests.get(url, params=params, timeout=30)
    _file_log(f"API GET {resp.url}")
    _file_log(f"  Status   : {resp.status_code}")
    _file_log(f"  Response : {resp.text}")

    parsed = None
    try:
        data = resp.json()
        parsed = data[0] if isinstance(data, list) else data
    except Exception:
        pass

    if parsed:
        msg = parsed.get("message", resp.text[:200])
        ok  = str(parsed.get("success", "")).lower() == "true"
    else:
        msg = resp.text[:200]
        ok  = 200 <= resp.status_code < 300

    log_fn(f"    → {resp.status_code}: {msg}\n")
    return ok, msg, parsed


# ─── Core Processing ──────────────────────────────────────────────────────────

def run_processing(epass_folder, gc_folder, db_cfg, ftp_cfg, api_cfg,
                   log_fn, progress_fn, done_fn):
    """
    Main processing logic (runs in a background thread):
      1. Read all PDFs from ePass and GC folders
      2. Extract passport + name from each
      3. Fetch listing API → contractor ID + which documents still needed
      4. Rename files in-place and upload via FTP
      5. Call API per worker for each uploaded document
      6. Return records for reporting
    """
    try:
        import pdfplumber
    except ImportError as e:
        done_fn(None, f"Missing library: {e}\n\nRun: pip install pdfplumber")
        return

    start = time.time()
    records = []
    errors  = []

    # ── Step 1: Collect all PDF files ────────────────────────────────────────
    epass_files = sorted([
        f for f in os.listdir(epass_folder) if f.lower().endswith(".pdf")
    ]) if os.path.isdir(epass_folder) else []

    gc_files = sorted([
        f for f in os.listdir(gc_folder) if f.lower().endswith(".pdf")
    ]) if os.path.isdir(gc_folder) else []

    log_fn(f"ePass folder: {len(epass_files)} file(s)\n")
    log_fn(f"GC folder:    {len(gc_files)} file(s)\n\n")

    # ── Step 2: Extract info from all PDFs ───────────────────────────────────
    log_fn("Reading PDFs...\n")

    epass_data = {}   # passport_no → {name, nationality, vp_no, epass_expiry, original_path}
    gc_data    = {}   # passport_no → {name, nationality, gc_expiry, original_path}

    for fname in epass_files:
        fpath = os.path.join(epass_folder, fname)
        try:
            with pdfplumber.open(fpath) as pdf:
                text = pdf.pages[0].extract_text() or ""
            pp, name, nat, vp, expiry = extract_epass_info(text)
            if not pp:
                log_fn(f"  WARNING: Could not extract passport from ePass: {fname}\n")
                errors.append({"file": fname, "type": "ePass", "error": "Passport not found"})
                continue
            epass_data[pp] = {
                "name":          name or "UNKNOWN",
                "nationality":   nat  or "",
                "vp_no":         vp   or "",
                "epass_expiry":  expiry or "",
                "original_path": fpath,
            }
            log_fn(f"  ePass: {pp} — {name}\n")
        except Exception as e:
            log_fn(f"  ERROR reading {fname}: {e}\n")
            errors.append({"file": fname, "type": "ePass", "error": str(e)})

    for fname in gc_files:
        fpath = os.path.join(gc_folder, fname)
        try:
            with pdfplumber.open(fpath) as pdf:
                text = pdf.pages[0].extract_text() or ""
            pp, name, nat, expiry = extract_gc_info(text)
            if not pp:
                log_fn(f"  WARNING: Could not extract passport from GC: {fname}\n")
                errors.append({"file": fname, "type": "GC", "error": "Passport not found"})
                continue
            gc_data[pp] = {
                "name":          name or "UNKNOWN",
                "nationality":   nat  or "",
                "gc_expiry":     expiry or "",
                "original_path": fpath,
            }
            log_fn(f"  GC:    {pp} — {name}\n")
        except Exception as e:
            log_fn(f"  ERROR reading {fname}: {e}\n")
            errors.append({"file": fname, "type": "GC", "error": str(e)})

    # ── Step 3: Build worker list (union of all passport numbers) ─────────────
    all_passports = sorted(set(list(epass_data.keys()) + list(gc_data.keys())))
    total = len(all_passports)

    if total == 0:
        done_fn(None, "No valid PDFs found in the selected folders.")
        return

    log_fn(f"\nFound {total} unique worker(s). Processing...\n\n")
    _file_log("=" * 60)
    _file_log(f"Session started — {total} worker(s)")

    # ── Step 4–6: Per-worker: rename → upload → API ───────────────────────────
    # Fetch listing to know which documents each worker still needs
    log_fn("Fetching pending worker list from API...\n")
    listing = fetch_listing(api_cfg["list_endpoint"], log_fn)

    log_fn("Connecting to FTP...\n")
    ftp = ftplib.FTP()
    ftp.connect(ftp_cfg["host"], ftp_cfg["port"], timeout=30)
    ftp.login(ftp_cfg["username"], ftp_cfg["password"])
    remote_root = ftp_cfg["remote_dir"].rstrip("/")
    _ftp_makedirs(ftp, remote_root, log_fn)
    log_fn(f"Connected. Remote root: {remote_root}/\n\n")

    for idx, passport_no in enumerate(all_passports, 1):
        ep   = epass_data.get(passport_no)
        gc   = gc_data.get(passport_no)

        # Prefer name from ePass, fall back to GC
        worker_name  = (ep or gc)["name"]
        nationality  = (ep or gc).get("nationality", "")

        # If PDF extraction missed the name, use the listing API's worker_name
        if worker_name == "UNKNOWN":
            listing_entry_pre = listing.get(passport_no) if listing else None
            if listing_entry_pre and listing_entry_pre.get("worker_name"):
                worker_name = listing_entry_pre["worker_name"]
        name_clean   = sanitize(worker_name)
        pp_clean     = sanitize(passport_no)

        log_fn(f"[{idx}/{total}] {passport_no} — {worker_name}\n")

        record = {
            "passport_no":      passport_no,
            "worker_name":      worker_name,
            "nationality":      nationality,
            "contractor_id":    None,
            "epass_path":       None,
            "gc_path":          None,
            "has_epass":        ep is not None,
            "has_gc":           gc is not None,
            "needs_epass":      None,   # set after listing lookup
            "needs_gc":         None,
            "status":           "Failed",
            "app_no":           "",
            "clab_id":          "",
            "api_timestamp":    "",
            "epass_api_result":  "",
            "gc_api_result":     "",
            "cont_email":        "",
            "email_status":      "",
            "epass_local_path":  ep["original_path"] if ep else None,
            "gc_local_path":     gc["original_path"] if gc else None,
            "error":             "",
        }

        # ── Resolve contractor ID + decide what to upload ────────────────────
        listing_entry = listing.get(passport_no) if listing is not None else None

        if listing is not None and not listing_entry:
            # Worker not in the pending list — all documents already set
            record["status"] = "Skipped - Already complete"
            log_fn(f"  Not in pending list — all documents already uploaded.\n\n")
            records.append(record)
            progress_fn(idx, total)
            continue

        if listing_entry:
            contractor_id = listing_entry["clab_id"]
            needs_epass   = listing_entry["needs_epass"]
            needs_gc      = listing_entry["needs_gc"]
        else:
            # Listing unavailable — fall back to DB
            contractor_id = db_lookup_contractor(passport_no, db_cfg)
            if not contractor_id:
                record["status"] = "Skipped - Not in DB"
                log_fn(f"  Passport not found in DB — skipping.\n\n")
                records.append(record)
                progress_fn(idx, total)
                continue
            needs_epass = ep is not None
            needs_gc    = gc is not None

        if not contractor_id:
            record["status"] = "Skipped - No CLAB ID"
            log_fn(f"  No CLAB ID found — skipping.\n\n")
            records.append(record)
            progress_fn(idx, total)
            continue

        record["contractor_id"] = contractor_id
        record["needs_epass"]   = needs_epass
        record["needs_gc"]      = needs_gc
        log_fn(f"  Contractor : {contractor_id}\n")
        log_fn(f"  Needs ePass: {'Yes' if needs_epass else 'No (already set)'}\n")
        log_fn(f"  Needs GC   : {'Yes' if needs_gc    else 'No (already set)'}\n")

        # ── Upload block (rename + FTP) ──────────────────────────────────────
        try:
            remote_company_dir = f"{remote_root}/{contractor_id}"
            epass_remote_path  = None
            gc_remote_path     = None

            _ftp_makedirs(ftp, remote_company_dir, log_fn)

            if ep and needs_epass:
                new_epass_name = f"{pp_clean} - {name_clean}_ePass.pdf"
                new_epass_path = os.path.join(epass_folder, new_epass_name)
                if ep["original_path"] != new_epass_path:
                    os.rename(ep["original_path"], new_epass_path)
                    log_fn(f"  Renamed ePass → {new_epass_name}\n")
                record["epass_local_path"] = new_epass_path
                remote_ep = f"{remote_company_dir}/{new_epass_name}"
                ftp_upload_file(ftp, new_epass_path, remote_ep, log_fn)
                epass_remote_path = remote_ep
                record["epass_path"] = remote_ep
            elif ep:
                log_fn(f"  ePass already set in DB — skipping upload.\n")

            if gc and needs_gc:
                new_gc_name = f"{pp_clean} - {name_clean}_GC.pdf"
                new_gc_path = os.path.join(gc_folder, new_gc_name)
                if gc["original_path"] != new_gc_path:
                    os.rename(gc["original_path"], new_gc_path)
                    log_fn(f"  Renamed GC    → {new_gc_name}\n")
                record["gc_local_path"] = new_gc_path
                remote_gc = f"{remote_company_dir}/{new_gc_name}"
                ftp_upload_file(ftp, new_gc_path, remote_gc, log_fn)
                gc_remote_path = remote_gc
                record["gc_path"] = remote_gc
            elif gc:
                log_fn(f"  GC already set in DB — skipping upload.\n")

            record["status"] = "Uploaded"
            log_fn(f"  Upload complete.\n")

        except Exception as e:
            record["error"] = str(e)
            _file_log(f"  UPLOAD FAILED [{passport_no}]: {e}")
            log_fn(f"  UPLOAD FAILED: {e}\n\n")
            records.append(record)
            progress_fn(idx, total)
            continue

        # ── API block ────────────────────────────────────────────────────────
        # FTP path  : /attach/CLAB002701/filename.pdf
        # API path  : ../attach/CLAB002701/filename.pdf
        def _api_path(ftp_path):
            return f"..{ftp_path}" if ftp_path else None

        log_fn(f"  Notifying API...\n")
        epass_ok = True
        gc_ok    = True

        # ePass API call
        if epass_remote_path:
            try:
                ok, msg, parsed = notify_api(
                    api_cfg["epass_endpoint"],
                    {"passno": passport_no, "wkrname": worker_name,
                     "clabid": contractor_id, "epass": _api_path(epass_remote_path)},
                    log_fn,
                )
                record["epass_api_result"] = "OK" if ok else f"Failed: {msg}"
                epass_ok = ok
                if parsed:
                    record["app_no"]        = record["app_no"]        or parsed.get("App_No",    "")
                    record["clab_id"]       = record["clab_id"]       or parsed.get("CLAB_ID",   "")
                    record["api_timestamp"] = record["api_timestamp"] or parsed.get("TimeStamp", "")
                    record["cont_email"]    = record["cont_email"]    or parsed.get("Cont_Email", "")
            except Exception as e:
                record["epass_api_result"] = f"Error: {e}"
                epass_ok = False
                _file_log(f"  ePass API ERROR [{passport_no}]: {e}")

        # GC / TP API call
        if gc_remote_path:
            try:
                ok, msg, parsed = notify_api(
                    api_cfg["gc_endpoint"],
                    {"passno": passport_no, "wkrname": worker_name,
                     "clabid": contractor_id, "tp": _api_path(gc_remote_path)},
                    log_fn,
                )
                record["gc_api_result"] = "OK" if ok else f"Failed: {msg}"
                gc_ok = ok
                if parsed:
                    record["app_no"]        = record["app_no"]        or parsed.get("App_No",    "")
                    record["clab_id"]       = record["clab_id"]       or parsed.get("CLAB_ID",   "")
                    record["api_timestamp"] = record["api_timestamp"] or parsed.get("TimeStamp", "")
                    record["cont_email"]    = record["cont_email"]    or parsed.get("Cont_Email", "")
            except Exception as e:
                record["gc_api_result"] = f"Error: {e}"
                gc_ok = False
                _file_log(f"  GC API ERROR [{passport_no}]: {e}")

        if epass_ok and gc_ok:
            record["status"] = "Success"
            log_fn(f"  Done.\n\n")
        else:
            record["status"] = "Uploaded / API Failed"
            log_fn(f"  API failure(s) — files uploaded successfully.\n\n")

        records.append(record)
        progress_fn(idx, total)

    if ftp:
        ftp.quit()

    log_fn("\nArchiving processed files...\n")
    archive_ts = move_processed_files(records, epass_folder, gc_folder, log_fn)

    elapsed     = time.time() - start
    n_success     = sum(1 for r in records if r["status"] == "Success")
    n_api_fail    = sum(1 for r in records if "API" in r["status"])
    n_upload_fail = sum(1 for r in records if r["status"] == "Failed")
    n_skipped     = sum(1 for r in records if r["status"].startswith("Skipped"))

    _file_log(f"Session complete — {total} workers, {n_skipped} skipped, "
              f"{n_upload_fail} upload failure(s), {n_api_fail} API failure(s). "
              f"Elapsed: {elapsed:.1f}s")

    stats = {
        "elapsed":      elapsed,
        "total":        total,
        "success":      n_success,
        "api_failures": n_api_fail,
        "failures":     n_upload_fail,
        "skipped":      n_skipped,
        "read_errors":  len(errors),
        "archive_ts":   archive_ts,
        "records":      records,
    }
    done_fn(stats, None)


# ─── File Archiving ───────────────────────────────────────────────────────────

def move_processed_files(records, epass_folder, gc_folder, log_fn):
    """
    Archive all processed PDFs into:
      Archive/{timestamp}/Done/ePass/   — successful ePass docs
      Archive/{timestamp}/Done/GC/      — successful GC docs
      Archive/{timestamp}/Failed/ePass/ — failed/unmatched ePass docs
      Archive/{timestamp}/Failed/GC/    — failed/unmatched GC docs
    Any remaining PDFs not tracked in records (read errors) are caught by a
    final folder scan and moved to Failed/ as well.
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    done_statuses = {"Success", "Uploaded / API Failed", "Skipped - Already complete"}

    # Pre-build all destination paths
    dest = {
        ("done", "epass"): os.path.join(ARCHIVE_DIR, ts, "Done",   "ePass"),
        ("done", "gc"):    os.path.join(ARCHIVE_DIR, ts, "Done",   "GC"),
        ("fail", "epass"): os.path.join(ARCHIVE_DIR, ts, "Failed", "ePass"),
        ("fail", "gc"):    os.path.join(ARCHIVE_DIR, ts, "Failed", "GC"),
    }
    for path in dest.values():
        os.makedirs(path, exist_ok=True)

    counts = {"done": 0, "fail": 0}

    def _move(src, outcome, doc_type):
        if not src or not os.path.isfile(src):
            return
        dst_dir  = dest[(outcome, doc_type)]
        dst_path = os.path.join(dst_dir, os.path.basename(src))
        if os.path.exists(dst_path):
            name, ext = os.path.splitext(os.path.basename(src))
            dst_path = os.path.join(dst_dir, f"{name}_{ts}{ext}")
        os.rename(src, dst_path)
        counts[outcome] += 1

    for r in records:
        outcome = "done" if r["status"] in done_statuses else "fail"
        if epass_folder and r.get("epass_local_path"):
            _move(r["epass_local_path"], outcome, "epass")
        if gc_folder and r.get("gc_local_path"):
            _move(r["gc_local_path"], outcome, "gc")

    # Sweep up remaining PDFs not tracked in records (e.g. read errors)
    for folder, doc_type in [(epass_folder, "epass"), (gc_folder, "gc")]:
        if not folder or not os.path.isdir(folder):
            continue
        for fname in os.listdir(folder):
            fpath = os.path.join(folder, fname)
            if fname.lower().endswith(".pdf") and os.path.isfile(fpath):
                _move(fpath, "fail", doc_type)

    log_fn(f"Archived to Archive/{ts}/  —  "
           f"{counts['done']} Done,  {counts['fail']} Failed\n")
    return ts


# ─── Email Notification ───────────────────────────────────────────────────────

def _build_email_html(app_no, workers):
    rows = ""
    for i, r in enumerate(workers):
        bg = "#f9f9f9" if i % 2 == 0 else "#ffffff"
        epass_cell = ("&#10003;" if r.get("epass_path") else
                      "Already set" if r.get("needs_epass") is False else "—")
        gc_cell    = ("&#10003;" if r.get("gc_path") else
                      "Already set" if r.get("needs_gc") is False else "—")
        rows += (
            f'<tr style="background:{bg};">'
            f'<td style="padding:8px 12px;">{r["worker_name"]}</td>'
            f'<td style="padding:8px 12px;">{r["passport_no"]}</td>'
            f'<td style="padding:8px 12px;text-align:center;">{epass_cell}</td>'
            f'<td style="padding:8px 12px;text-align:center;">{gc_cell}</td>'
            f'</tr>'
        )
    return f"""\
<html>
<body style="font-family:Arial,sans-serif;color:#333;background:#f5f5f5;padding:20px;">
  <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:8px;
              overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
    <div style="background:#0a3d62;color:#fff;padding:24px;">
      <h2 style="margin:0;">Document Submission Notice</h2>
      <p style="margin:4px 0 0;opacity:0.8;">Construction Labour Exchange Centre Berhad</p>
    </div>
    <div style="padding:24px;">
      <p>Dear Contractor,</p>
      <p>The following worker document(s) have been submitted for
         <strong>Application No: {app_no}</strong>:</p>
      <table style="width:100%;border-collapse:collapse;margin:16px 0;">
        <thead>
          <tr style="background:#0a3d62;color:#fff;">
            <th style="padding:10px 12px;text-align:left;">Worker Name</th>
            <th style="padding:10px 12px;text-align:left;">Passport No</th>
            <th style="padding:10px 12px;text-align:center;">ePass</th>
            <th style="padding:10px 12px;text-align:center;">GC / TP</th>
          </tr>
        </thead>
        <tbody>{rows}</tbody>
      </table>
      <p style="color:#666;font-size:13px;">
        This is an automated notification. Please do not reply to this email.
      </p>
    </div>
    <div style="background:#f0f2f5;padding:16px;text-align:center;
                font-size:12px;color:#888;">
      Construction Labour Exchange Centre Berhad (CLAB)
    </div>
  </div>
</body>
</html>"""


def send_client_emails(records, log_fn):
    """Send one email per unique App_No, skipping duplicates and missing emails."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from collections import defaultdict

    try:
        email_cfg = load_email_config()
    except Exception as e:
        log_fn(f"  Email skipped — could not load config: {e}\n")
        return

    if not email_cfg["username"] or not email_cfg["sender_email"]:
        log_fn("  Email skipped — SMTP credentials not configured in config.ini.\n")
        return

    # Group successful records by App_No; skip records with no app_no or no email
    by_app = defaultdict(list)
    for r in records:
        app_no     = r.get("app_no", "").strip()
        cont_email = r.get("cont_email", "").strip()
        if r["status"] == "Success":
            if not app_no:
                r["email_status"] = "No App No"
            elif not cont_email:
                r["email_status"] = "No email"
            else:
                by_app[app_no].append(r)

    if not by_app:
        log_fn("  Email skipped — no successful records with App No and contractor email.\n")
        return

    log_fn(f"\nSending email notifications ({len(by_app)} application(s))...\n")
    try:
        smtp = smtplib.SMTP(email_cfg["smtp_host"], email_cfg["smtp_port"], timeout=30)
        if email_cfg["use_tls"]:
            smtp.starttls()
        smtp.login(email_cfg["username"], email_cfg["password"])

        for app_no, workers in by_app.items():
            cont_email = workers[0]["cont_email"].strip()
            try:
                msg = MIMEMultipart("alternative")
                msg["Subject"] = f"Document Submission — Application {app_no}"
                msg["From"]    = f"{email_cfg['sender_name']} <{email_cfg['sender_email']}>"
                msg["To"]      = cont_email
                msg.attach(MIMEText(_build_email_html(app_no, workers), "html"))
                smtp.sendmail(email_cfg["sender_email"], cont_email, msg.as_string())
                for r in workers:
                    r["email_status"] = f"Sent → {cont_email}"
                log_fn(f"  Email sent → {cont_email}  (App No: {app_no}, {len(workers)} worker(s))\n")
                _file_log(f"Email sent → {cont_email} App No: {app_no}")
            except Exception as e:
                for r in workers:
                    r["email_status"] = f"Failed: {e}"
                log_fn(f"  Email FAILED for App No {app_no}: {e}\n")
                _file_log(f"Email FAILED App No: {app_no} — {e}")

        smtp.quit()
    except Exception as e:
        log_fn(f"  Email ERROR: {e}\n")
        _file_log(f"Email ERROR: {e}")


# ─── Excel Report ─────────────────────────────────────────────────────────────

def generate_report(records, stats=None):
    """
    Generate a three-sheet Excel report:
      Sheet 1 — Run Summary      (statistics for the run)
      Sheet 2 — Upload Summary   (one row per worker)
    Returns the saved report file path.
    """
    stats = stats or {}

    HDR_FILL  = PatternFill("solid", fgColor="0A3D62")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    OK_FILL   = PatternFill("solid", fgColor="E8F5E9")
    FAIL_FILL = PatternFill("solid", fgColor="FFEBEE")
    ALT_FILL  = PatternFill("solid", fgColor="F5F8FA")
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    THIN      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _set_header(ws, headers, col_widths):
        ws.row_dimensions[1].height = 20
        for col, (title, width) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(1, col, title)
            c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

    def _style_row(ws, row_idx, n_cols, fill):
        for col in range(1, n_cols + 1):
            c = ws.cell(row_idx, col)
            c.fill   = fill
            c.border = BORDER
            c.alignment = CENTER if col == 1 else LEFT

    wb = openpyxl.Workbook()

    # ── Sheet 1: Run Summary ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Run Summary"
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 35

    SECT_FILL = PatternFill("solid", fgColor="1A5276")
    SECT_FONT = Font(bold=True, color="FFFFFF", size=10)
    LBL_FILL  = PatternFill("solid", fgColor="F2F3F4")
    LBL_FONT  = Font(color="444444", size=10)
    VAL_FONT  = Font(bold=True, size=10)
    INDENT    = Alignment(horizontal="left", vertical="center", indent=1)

    def _sum_title(row, text):
        ws_sum.merge_cells(f"A{row}:B{row}")
        c = ws_sum.cell(row, 1, text)
        c.fill      = HDR_FILL
        c.font      = Font(bold=True, color="FFFFFF", size=14)
        c.alignment = INDENT
        ws_sum.row_dimensions[row].height = 30

    def _sum_section(row, text):
        ws_sum.merge_cells(f"A{row}:B{row}")
        c = ws_sum.cell(row, 1, text)
        c.fill      = SECT_FILL
        c.font      = SECT_FONT
        c.alignment = INDENT
        ws_sum.row_dimensions[row].height = 18

    def _sum_row(row, label, value, ok=None):
        lc = ws_sum.cell(row, 1, label)
        vc = ws_sum.cell(row, 2, value)
        lc.fill, lc.font, lc.alignment = LBL_FILL, LBL_FONT, INDENT
        vc.font, vc.alignment = VAL_FONT, INDENT
        if ok is True:
            vc.fill = OK_FILL
        elif ok is False:
            vc.fill = FAIL_FILL
        ws_sum.row_dimensions[row].height = 16

    def _sum_gap(row):
        ws_sum.row_dimensions[row].height = 6

    # Compute values
    n_total    = len(records)
    n_success  = sum(1 for r in records if r["status"] == "Success")
    n_complete = sum(1 for r in records if r["status"] == "Skipped - Already complete")
    n_no_db    = sum(1 for r in records if r["status"] == "Skipped - Not in DB")
    n_no_clab  = sum(1 for r in records if r["status"] == "Skipped - No CLAB ID")
    n_api_fail = sum(1 for r in records if "API" in r["status"] and r["status"] != "Success")
    n_up_fail  = sum(1 for r in records if r["status"] == "Failed")
    n_epass    = sum(1 for r in records if r.get("epass_path"))
    n_gc       = sum(1 for r in records if r.get("gc_path"))
    sent_apps  = {r.get("app_no") for r in records
                  if r.get("email_status", "").startswith("Sent") and r.get("app_no")}
    fail_apps  = {r.get("app_no") for r in records
                  if r.get("email_status", "").startswith("Failed") and r.get("app_no")}
    n_email_na = sum(1 for r in records
                     if r.get("email_status") in ("No App No", "No email", ""))

    elapsed     = stats.get("elapsed", 0)
    read_errors = stats.get("read_errors", 0)
    archive_ts  = stats.get("archive_ts", "")
    generated   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    r = 1
    _sum_title(r,   "Run Summary");                                            r += 1
    _sum_gap(r);                                                               r += 1
    _sum_section(r, "Run Information");                                        r += 1
    _sum_row(r, "Generated",       generated);                                 r += 1
    _sum_row(r, "Elapsed Time",    f"{elapsed:.1f} seconds");                  r += 1
    _sum_row(r, "Archive Folder",  f"Archive/{archive_ts}" if archive_ts
                                   else "N/A");                                r += 1
    _sum_gap(r);                                                               r += 1
    _sum_section(r, "Processing Results");                                     r += 1
    _sum_row(r, "Total Workers",         n_total);                             r += 1
    _sum_row(r, "Success",               n_success,
             ok=True if n_success > 0 else None);                              r += 1
    _sum_row(r, "Already Complete",      n_complete);                          r += 1
    _sum_row(r, "Upload Failed",         n_up_fail,
             ok=False if n_up_fail > 0 else None);                             r += 1
    _sum_row(r, "API Failed",            n_api_fail,
             ok=False if n_api_fail > 0 else None);                            r += 1
    _sum_row(r, "Skipped (Not in DB)",   n_no_db,
             ok=False if n_no_db > 0 else None);                               r += 1
    _sum_row(r, "Skipped (No CLAB ID)",  n_no_clab,
             ok=False if n_no_clab > 0 else None);                             r += 1
    _sum_row(r, "Read Errors",           read_errors,
             ok=False if read_errors > 0 else None);                           r += 1
    _sum_gap(r);                                                               r += 1
    _sum_section(r, "Documents Uploaded");                                     r += 1
    _sum_row(r, "ePass",                 n_epass);                             r += 1
    _sum_row(r, "GC / TP",              n_gc);                                 r += 1
    _sum_gap(r);                                                               r += 1
    _sum_section(r, "Email Notifications");                                    r += 1
    _sum_row(r, "Applications Emailed",  len(sent_apps),
             ok=True if sent_apps else None);                                  r += 1
    _sum_row(r, "Email Failed",          len(fail_apps),
             ok=False if fail_apps else None);                                 r += 1
    _sum_row(r, "No Email / App No",     n_email_na);                          r += 1

    # ── Sheet 2: Upload Summary ───────────────────────────────────────────────
    ws1 = wb.create_sheet("Upload Summary")
    ws1.freeze_panes = "A2"

    hdrs1   = ["#", "Passport No", "Worker Name", "Nationality", "Contractor ID",
               "ePass", "GC", "Status", "App No", "CLAB ID", "Timestamp",
               "ePass API", "GC API", "Email", "Error"]
    widths1 = [5,   15,            30,             15,            15,
               8,    8,    16,      14,       14,        20,
               30,         30,       30,      40]
    _set_header(ws1, hdrs1, widths1)

    def _doc_col(has_file, needed, path):
        """Summarise document status for one worker/doc-type cell."""
        if not has_file:
            return "No file"
        if needed is None:
            return "In folder"
        if not needed:
            return "Already set"
        return "Uploaded" if path else "Upload failed"

    for i, r in enumerate(records, 1):
        row_idx = i + 1
        alt     = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        if r["status"] == "Success":
            fill = OK_FILL
        elif r["status"].startswith("Skipped"):
            fill = ALT_FILL
        elif r["error"] or "Failed" in r["status"] or "Error" in r["status"]:
            fill = FAIL_FILL
        else:
            fill = alt

        epass_api = ("Not needed" if r.get("needs_epass") is False
                     else r.get("epass_api_result") or "N/A")
        gc_api    = ("Not needed" if r.get("needs_gc")    is False
                     else r.get("gc_api_result")    or "N/A")

        ws1.append([
            i,
            r["passport_no"],
            r["worker_name"],
            r["nationality"],
            r["contractor_id"] or "",
            _doc_col(r["has_epass"], r.get("needs_epass"), r["epass_path"]),
            _doc_col(r["has_gc"],    r.get("needs_gc"),    r["gc_path"]),
            r["status"],
            r.get("app_no",        ""),
            r.get("clab_id",       ""),
            r.get("api_timestamp", ""),
            epass_api,
            gc_api,
            r.get("email_status",  ""),
            r["error"],
        ])
        _style_row(ws1, row_idx, len(hdrs1), fill)

    # Totals row
    total_row = len(records) + 2
    ws1.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    n_ok   = sum(1 for r in records if r["status"] == "Success")
    n_skip = sum(1 for r in records if r["status"].startswith("Skipped"))
    n_fail = len(records) - n_ok - n_skip
    ws1.cell(total_row, 8,
             f"{n_ok} Success / {n_skip} Skipped / {n_fail} Failed"
             ).font = Font(bold=True)

    os.makedirs(REPORT_DIR, exist_ok=True)
    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(REPORT_DIR, f"EPassReport_{ts}.xlsx")
    wb.save(report_path)
    return report_path


# ─── Headless / Scheduled Entry Point ────────────────────────────────────────

def _run_headless():
    """Run processing without the GUI (invoked via --scheduled flag)."""
    _file_log("=" * 60)
    _file_log("Scheduled run started")

    try:
        sched       = load_schedule_config()
        epass_folder = sched["epass_folder"]
        gc_folder    = sched["gc_folder"]
        if not epass_folder and not gc_folder:
            _file_log("ERROR: No folders configured in [Schedule] section of config.ini")
            sys.exit(1)
        db_cfg  = load_db_config()
        ftp_cfg = load_ftp_config()
        api_cfg = load_api_config()
    except Exception as e:
        _file_log(f"ERROR loading config: {e}")
        sys.exit(1)

    def log_fn(msg):
        for line in msg.split("\n"):
            line = line.strip()
            if line:
                _file_log(line)

    result = {}

    def done_fn(stats, error):
        result["stats"] = stats
        result["error"] = error

    run_processing(epass_folder, gc_folder, db_cfg, ftp_cfg, api_cfg,
                   log_fn, lambda c, t: None, done_fn)

    if result.get("error"):
        _file_log(f"ERROR: {result['error']}")
        sys.exit(1)

    stats = result.get("stats") or {}
    if stats.get("records"):
        send_client_emails(stats["records"], log_fn)
        try:
            report_path = generate_report(stats["records"], stats)
            _file_log(f"Report saved: {report_path}")
        except Exception as e:
            _file_log(f"Report error: {e}")

    _file_log(
        f"Done — {stats.get('total', 0)} workers | "
        f"{stats.get('success', 0)} success | "
        f"{stats.get('skipped', 0)} skipped | "
        f"{stats.get('api_failures', 0)} API failed | "
        f"{stats.get('failures', 0)} upload failed"
    )
    _file_log("=" * 60)


# ─── Scheduler Dialog ─────────────────────────────────────────────────────────

class SchedulerDialog(tk.Toplevel):
    TASK_NAME = "ePass TP Upload"

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Schedule Daily Run")
        self.resizable(False, False)
        self.grab_set()
        self.configure(bg="#f0f2f5")

        sched = load_schedule_config()
        self._epass_var  = tk.StringVar(value=sched["epass_folder"])
        self._gc_var     = tk.StringVar(value=sched["gc_folder"])
        self._time_var   = tk.StringVar(value=sched["run_time"])
        self._status_var = tk.StringVar()
        self._toggle_btn = None   # Disable/Enable button ref

        self._build()
        self._refresh_status()
        self._load_history()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        w, h = 620, 520
        px = parent.winfo_rootx() + parent.winfo_width()  // 2
        py = parent.winfo_rooty() + parent.winfo_height() // 2
        self.geometry(f"{w}x{h}+{px - w // 2}+{py - h // 2}")

    def _build(self):
        pad = {"padx": 16, "pady": 5}

        # ── Run time ──
        time_row = tk.Frame(self, bg="#f0f2f5")
        time_row.pack(fill="x", **pad)
        tk.Label(time_row, text="Run time (HH:MM):", bg="#f0f2f5",
                 font=("Inter", 10)).pack(side="left")
        tk.Entry(time_row, textvariable=self._time_var, width=8,
                 font=("Inter", 10)).pack(side="left", padx=(8, 0))

        # ── ePass folder ──
        tk.Label(self, text="ePass folder:", bg="#f0f2f5",
                 font=("Inter", 10)).pack(anchor="w", padx=16, pady=(4, 0))
        ef = tk.Frame(self, bg="#f0f2f5")
        ef.pack(fill="x", padx=16)
        tk.Entry(ef, textvariable=self._epass_var, width=60,
                 font=("Inter", 9)).pack(side="left")
        ttk.Button(ef, text="Browse...",
                   command=lambda: self._browse(self._epass_var)
                   ).pack(side="left", padx=(6, 0))

        # ── GC folder ──
        tk.Label(self, text="GC folder:", bg="#f0f2f5",
                 font=("Inter", 10)).pack(anchor="w", padx=16, pady=(4, 0))
        gf = tk.Frame(self, bg="#f0f2f5")
        gf.pack(fill="x", padx=16)
        tk.Entry(gf, textvariable=self._gc_var, width=60,
                 font=("Inter", 9)).pack(side="left")
        ttk.Button(gf, text="Browse...",
                   command=lambda: self._browse(self._gc_var)
                   ).pack(side="left", padx=(6, 0))

        # ── Status label ──
        tk.Label(self, textvariable=self._status_var, bg="#f0f2f5",
                 font=("Inter", 9), fg="#555").pack(anchor="w", padx=16, pady=(8, 0))

        # ── Action buttons ──
        btn_row = tk.Frame(self, bg="#f0f2f5")
        btn_row.pack(fill="x", padx=16, pady=(6, 6))
        ttk.Button(btn_row, text="Register Task",
                   command=self._register_task).pack(side="left")
        self._toggle_btn = ttk.Button(btn_row, text="Disable Task",
                                      command=self._toggle_task)
        self._toggle_btn.pack(side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Remove Task",
                   command=self._remove_task).pack(side="left", padx=(8, 0))
        ttk.Button(btn_row, text="Close",
                   command=self.destroy).pack(side="right")

        # ── History section ──
        sep = tk.Frame(self, bg="#ccd3db", height=1)
        sep.pack(fill="x", padx=16, pady=(4, 0))
        tk.Label(self, text="Schedule Run History", bg="#f0f2f5",
                 font=("Inter", 9, "bold")).pack(anchor="w", padx=16, pady=(6, 2))

        tree_frame = tk.Frame(self, bg="#f0f2f5")
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        cols = ("date", "workers", "success", "skipped", "api_failed", "upload_failed")
        self._tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=8)

        headings = {
            "date":          ("Date / Time",    160),
            "workers":       ("Workers",          60),
            "success":       ("Success",          60),
            "skipped":       ("Skipped",          60),
            "api_failed":    ("API Failed",       70),
            "upload_failed": ("Upload Failed",    80),
        }
        for col, (text, width) in headings.items():
            self._tree.heading(col, text=text)
            self._tree.column(col, width=width, anchor="center")

        sb = ttk.Scrollbar(tree_frame, orient="vertical",
                           command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # row colouring
        self._tree.tag_configure("ok",   background="#E8F5E9")
        self._tree.tag_configure("fail", background="#FFEBEE")
        self._tree.tag_configure("alt",  background="#F5F8FA")

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _browse(self, var):
        path = filedialog.askdirectory(parent=self)
        if path:
            var.set(path)

    def _get_task_status(self):
        """Returns 'enabled', 'disabled', or None if the task doesn't exist."""
        r = subprocess.run(
            ["schtasks", "/query", "/tn", self.TASK_NAME, "/fo", "LIST"],
            capture_output=True, text=True
        )
        if r.returncode != 0:
            return None
        for line in r.stdout.splitlines():
            if line.strip().lower().startswith("status:"):
                val = line.split(":", 1)[1].strip().lower()
                return "disabled" if "disabled" in val else "enabled"
        return "enabled"

    def _refresh_status(self):
        status = self._get_task_status()
        if status is None:
            self._status_var.set("Status: No scheduled task registered.")
            if self._toggle_btn:
                self._toggle_btn.config(state="disabled", text="Disable Task")
        elif status == "disabled":
            self._status_var.set(
                f"Status: Task registered — DISABLED (runs daily at {self._time_var.get()})")
            if self._toggle_btn:
                self._toggle_btn.config(state="normal", text="Enable Task")
        else:
            self._status_var.set(
                f"Status: Task registered — ENABLED (runs daily at {self._time_var.get()})")
            if self._toggle_btn:
                self._toggle_btn.config(state="normal", text="Disable Task")

    def _load_history(self):
        """Parse upload.log and populate the history treeview."""
        self._tree.delete(*self._tree.get_children())
        if not os.path.exists(LOG_FILE_PATH):
            return

        with open(LOG_FILE_PATH, "r", encoding="utf-8") as f:
            lines = f.readlines()

        entries = []
        for i, line in enumerate(lines):
            if "Scheduled run started" not in line:
                continue
            m = re.match(r"\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]", line)
            timestamp = m.group(1) if m else "Unknown"
            for j in range(i + 1, min(i + 300, len(lines))):
                done = lines[j]
                if "Done —" not in done:
                    continue
                dm = re.search(
                    r"(\d+) workers.*?(\d+) success.*?(\d+) skipped.*?"
                    r"(\d+) API failed.*?(\d+) upload failed",
                    done
                )
                if dm:
                    entries.append((timestamp,) + dm.groups())
                break

        for idx, (ts, workers, success, skipped, api_fail, up_fail) in \
                enumerate(reversed(entries)):
            tag = ("ok"  if int(api_fail) == 0 and int(up_fail) == 0 else
                   "fail" if (int(api_fail) + int(up_fail)) == int(workers) else
                   "alt")
            self._tree.insert("", "end",
                              values=(ts, workers, success, skipped, api_fail, up_fail),
                              tags=(tag,))

    # ── Actions ──────────────────────────────────────────────────────────────

    def _register_task(self):
        run_time = self._time_var.get().strip()
        if not re.match(r"^\d{2}:\d{2}$", run_time):
            messagebox.showerror("Invalid time",
                "Enter time as HH:MM (e.g. 08:00).", parent=self)
            return
        epass_folder = self._epass_var.get().strip()
        gc_folder    = self._gc_var.get().strip()
        if not epass_folder and not gc_folder:
            messagebox.showerror("No folders",
                "Set at least one folder path.", parent=self)
            return

        save_schedule_config(epass_folder, gc_folder, run_time)

        if getattr(sys, 'frozen', False):
            tr = f'"{sys.executable}" --scheduled'
        else:
            tr = f'"{sys.executable}" "{os.path.abspath(__file__)}" --scheduled'

        try:
            subprocess.run(
                ["schtasks", "/create",
                 "/tn", self.TASK_NAME,
                 "/tr", tr,
                 "/sc", "daily",
                 "/st", run_time,
                 "/f"],
                check=True, capture_output=True
            )
            self._refresh_status()
            messagebox.showinfo("Registered",
                f"Task registered.\nRuns daily at {run_time}.", parent=self)
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error",
                f"Failed to register task:\n{e.stderr.decode()}", parent=self)

    def _toggle_task(self):
        status = self._get_task_status()
        if status is None:
            messagebox.showinfo("Not found", "No task registered.", parent=self)
            return
        action = "/disable" if status == "enabled" else "/enable"
        label  = "disable" if status == "enabled" else "enable"
        try:
            subprocess.run(
                ["schtasks", "/change", "/tn", self.TASK_NAME, action],
                check=True, capture_output=True
            )
            self._refresh_status()
            messagebox.showinfo("Done",
                f"Task {label}d successfully.", parent=self)
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error",
                f"Failed to {label} task:\n{e.stderr.decode()}", parent=self)

    def _remove_task(self):
        if self._get_task_status() is None:
            messagebox.showinfo("Not found", "No task to remove.", parent=self)
            return
        if not messagebox.askyesno("Confirm",
                "Remove the scheduled task permanently?", parent=self):
            return
        try:
            subprocess.run(
                ["schtasks", "/delete", "/tn", self.TASK_NAME, "/f"],
                check=True, capture_output=True
            )
            self._refresh_status()
            messagebox.showinfo("Removed", "Scheduled task removed.", parent=self)
        except subprocess.CalledProcessError as e:
            messagebox.showerror("Error",
                f"Failed to remove task:\n{e.stderr.decode()}", parent=self)


# ─── Login Window ─────────────────────────────────────────────────────────────

class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ePass TP Upload — Login")
        self.resizable(False, False)
        self.result = False
        self._attempts = 0

        try:
            self.iconbitmap(resource_path('favicon.ico'))
        except Exception:
            pass

        dpi = self.winfo_fpixels('1i')
        self.tk.call('tk', 'scaling', dpi / 72.0)

        self._build_ui()
        self._center_window()
        self.bind("<Return>", lambda e: self._attempt_login())
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _center_window(self):
        self.update_idletasks()
        w, h = 380, 340
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TButton", padding=6, font=("Inter", 10))
        style.configure("Login.TButton",
                        background="#0a3d62", foreground="white",
                        font=("Inter", 11, "bold"), padding=10)
        style.map("Login.TButton",
                  background=[("active", "#0c4a75")])

        # Header
        hdr = tk.Frame(self, bg="#0a3d62", pady=20)
        hdr.pack(fill="x")
        tk.Label(hdr, text="ePass TP Upload",
                 bg="#0a3d62", fg="white",
                 font=("Inter", 16, "bold")).pack()
        tk.Label(hdr, text="Construction Labour Exchange Centre Berhad",
                 bg="#0a3d62", fg="#a8c5da",
                 font=("Inter", 9)).pack()

        # Form
        form = tk.Frame(self, bg="#f0f2f5", padx=30, pady=20)
        form.pack(fill="both", expand=True)

        tk.Label(form, text="Username", bg="#f0f2f5",
                 font=("Inter", 10)).pack(anchor="w")
        self._user_var = tk.StringVar()
        self._user_entry = tk.Entry(form, textvariable=self._user_var,
                                    font=("Inter", 11), width=28)
        self._user_entry.pack(fill="x", pady=(2, 12))

        tk.Label(form, text="Password", bg="#f0f2f5",
                 font=("Inter", 10)).pack(anchor="w")
        self._pass_var = tk.StringVar()
        tk.Entry(form, textvariable=self._pass_var, show="•",
                 font=("Inter", 11), width=28).pack(fill="x", pady=(2, 4))

        self._error_var = tk.StringVar()
        tk.Label(form, textvariable=self._error_var,
                 bg="#f0f2f5", fg="#c0392b",
                 font=("Inter", 9)).pack(anchor="w", pady=(0, 10))

        ttk.Button(form, text="Login", style="Login.TButton",
                   command=self._attempt_login).pack(fill="x")

        self._user_entry.focus_set()

    def _attempt_login(self):
        username = self._user_var.get().strip()
        password = self._pass_var.get()

        if not username:
            self._error_var.set("Please enter your username.")
            return

        try:
            db_cfg = load_login_db_config()
            conn = pymysql.connect(
                host=db_cfg["host"],
                port=db_cfg["port"],
                user=db_cfg["username"],
                password=db_cfg["password"],
                database=db_cfg["database"],
                connect_timeout=10,
                charset="utf8mb4",
            )
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT 1 FROM user WHERE username = %s AND password = %s LIMIT 1",
                        (username, password),
                    )
                    found = cur.fetchone() is not None
            finally:
                conn.close()
        except Exception as e:
            messagebox.showerror("Database Error",
                                 f"Could not connect to database:\n{e}", parent=self)
            return

        if found:
            self.result = True
            self.destroy()
        else:
            self._attempts += 1
            self._error_var.set(
                f"Invalid username or password.  (Attempt {self._attempts})")
            self._pass_var.set("")


# ─── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ePass TP Upload")
        self.resizable(False, False)
        self._running       = False
        self._records       = []
        self._report_path   = None
        _ensure_config()

        try:
            self.iconbitmap(resource_path('favicon.ico'))
        except Exception:
            pass

        dpi = self.winfo_fpixels('1i')
        self._scale = dpi / 96.0
        self.tk.call('tk', 'scaling', dpi / 72.0)

        self._build_ui()
        self._center_window()

    def _center_window(self):
        self.update_idletasks()
        w = int(680 * self._scale)
        h = int(580 * self._scale)
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

    def _build_ui(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TButton", padding=6, font=("Inter", 10))
        style.configure("Primary.TButton",
                        background="#0a3d62", foreground="white",
                        font=("Inter", 11, "bold"), padding=10)
        style.map("Primary.TButton",
                  background=[("active", "#0c4a75"), ("disabled", "#cccccc")])

        # ── Header ──
        hdr = tk.Frame(self, bg="#0a3d62", pady=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="ePass TP Upload",
                 bg="#0a3d62", fg="white",
                 font=("Inter", 16, "bold")).pack()
        tk.Label(hdr, text="Construction Labour Exchange Centre Berhad",
                 bg="#0a3d62", fg="#a8c5da",
                 font=("Inter", 9)).pack()

        # ── Folder pickers ──
        picker = tk.Frame(self, bg="#f0f2f5", pady=12, padx=20)
        picker.pack(fill="x")

        self._epass_var = tk.StringVar(value="No folder selected")
        self._gc_var    = tk.StringVar(value="No folder selected")

        for label_text, str_var, attr in [
            ("ePass Folder:", self._epass_var, "_epass_folder"),
            ("GC Folder:",    self._gc_var,    "_gc_folder"),
        ]:
            tk.Label(picker, text=label_text, bg="#f0f2f5",
                     font=("Inter", 10)).pack(anchor="w", pady=(6, 0))
            row = tk.Frame(picker, bg="#f0f2f5")
            row.pack(fill="x", pady=(2, 0))
            tk.Label(row, textvariable=str_var, bg="white",
                     relief="sunken", anchor="w", padx=8,
                     font=("Inter", 9), width=55, fg="#555").pack(side="left", ipady=5)
            setattr(self, attr, None)
            ttk.Button(row, text="Browse...",
                       command=lambda a=attr, v=str_var: self._browse_folder(a, v)
                       ).pack(side="left", padx=(8, 0))


        # ── Process button ──
        btn_frame = tk.Frame(self, bg="#f0f2f5", pady=8)
        btn_frame.pack()
        self._process_btn = ttk.Button(btn_frame, text="Process & Upload",
                                       style="Primary.TButton",
                                       command=self._start_processing,
                                       state="disabled")
        self._process_btn.pack(ipadx=20)

        # ── Progress bar ──
        prog_frame = tk.Frame(self, bg="#f0f2f5", padx=20)
        prog_frame.pack(fill="x")
        self._progress = ttk.Progressbar(prog_frame, mode="determinate", maximum=100)
        self._progress.pack(fill="x", pady=(0, 4))

        # ── Post-process action buttons ──
        self._post_frame = tk.Frame(self, bg="#f0f2f5", pady=6)
        self._post_frame.pack(fill="x", padx=20)

        self._report_btn = ttk.Button(self._post_frame, text="Open Excel Report",
                                      command=self._open_report, state="disabled")
        self._report_btn.pack(side="left")

        ttk.Button(self._post_frame, text="Schedule...",
                   command=self._open_scheduler).pack(side="right")

        # ── Log ──
        log_frame = tk.Frame(self, padx=20, pady=8, bg="#f0f2f5")
        log_frame.pack(fill="both", expand=True)
        tk.Label(log_frame, text="Log:", bg="#f0f2f5",
                 font=("Inter", 9, "bold"), anchor="w").pack(anchor="w")
        self._log = scrolledtext.ScrolledText(
            log_frame, state="disabled", height=12,
            font=("Consolas", 9), bg="#1e1e1e", fg="#d4d4d4",
            insertbackground="white", relief="flat")
        self._log.pack(fill="both", expand=True)

        # ── Status bar ──
        self._status_var = tk.StringVar(value="Ready — select ePass and GC folders to begin.")
        sb = tk.Label(self, textvariable=self._status_var,
                      bg="#dde3ea", fg="#333",
                      font=("Inter", 9), anchor="w", padx=12, pady=4)
        sb.pack(fill="x", side="bottom")

    def _browse_folder(self, attr, str_var):
        path = filedialog.askdirectory(title=f"Select {attr.replace('_', ' ').title()}")
        if path:
            setattr(self, attr, path)
            str_var.set(path)
            self._check_ready()

    def _check_ready(self):
        """Enable Process button only when at least one folder is selected."""
        if self._epass_folder or self._gc_folder:
            self._process_btn.config(state="normal")
            self._status_var.set("Ready — click 'Process & Upload' to begin.")
        else:
            self._process_btn.config(state="disabled")

    def _log_msg(self, msg: str):
        def _append():
            self._log.config(state="normal")
            self._log.insert("end", msg)
            self._log.see("end")
            self._log.config(state="disabled")
        self.after(0, _append)

    def _progress_update(self, current, total):
        def _update():
            self._progress['value'] = (current / total) * 100
            self._status_var.set(f"Processing... {current}/{total} workers")
        self.after(0, _update)

    def _start_processing(self):
        if self._running:
            return

        try:
            db_cfg  = load_db_config()
            ftp_cfg = load_ftp_config()
            api_cfg = load_api_config()
        except Exception as e:
            messagebox.showerror("Config Error", str(e))
            return

        epass_folder = self._epass_folder or ""
        gc_folder    = self._gc_folder    or ""

        if not epass_folder and not gc_folder:
            messagebox.showwarning("No Folders", "Please select at least one folder.")
            return

        self._running = True
        self._process_btn.config(state="disabled")
        self._report_btn.config(state="disabled")
        self._progress['value'] = 0
        self._log_msg(f"{'─' * 60}\n")
        self._status_var.set("Processing — please wait.")

        threading.Thread(
            target=run_processing,
            args=(epass_folder, gc_folder, db_cfg, ftp_cfg, api_cfg,
                  self._log_msg, self._progress_update, self._on_done),
            daemon=True,
        ).start()

    def _on_done(self, stats, error):
        def _update():
            self._running = False
            self._process_btn.config(state="normal")

            if error:
                self._progress['value'] = 0
                self._log_msg(f"\nERROR: {error}\n")
                self._status_var.set("Failed — see log for details.")
                messagebox.showerror("Error", error)
                return

            self._records = stats.get("records", [])
            self._progress['value'] = 100
            self._log_msg(
                f"\nCompleted in {stats['elapsed']:.1f}s\n"
                f"  Workers:        {stats['total']}\n"
                f"  Success:        {stats['success']}\n"
                f"  Skipped:        {stats['skipped']}\n"
                f"  API failures:   {stats['api_failures']}\n"
                f"  Upload failures:{stats['failures']}\n"
                f"  Read errors:    {stats['read_errors']}\n"
            )
            self._status_var.set(
                f"Done — {stats['total']} workers | "
                f"{stats['success']} success | "
                f"{stats['skipped']} skipped | "
                f"{stats['api_failures']} API failed | "
                f"{stats['failures']} upload failed."
            )

            # Email notifications then report
            if self._records:
                send_client_emails(self._records, self._log_msg)
                try:
                    self._report_path = generate_report(self._records, stats)
                    self._log_msg(f"Report saved: {os.path.basename(self._report_path)}\n")
                    self._report_btn.config(state="normal")
                except Exception as e:
                    self._log_msg(f"Report error: {e}\n")

        self.after(0, _update)

    def _open_report(self):
        if self._report_path and os.path.exists(self._report_path):
            os.startfile(self._report_path)
        else:
            messagebox.showinfo("Info", "Report not found.")

    def _open_scheduler(self):
        SchedulerDialog(self)


if __name__ == "__main__":
    if "--scheduled" in sys.argv:
        _run_headless()
    else:
        login = LoginWindow()
        login.mainloop()
        if login.result:
            app = App()
            app.mainloop()
